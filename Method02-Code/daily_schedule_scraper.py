"""
daily_schedule_scraper.py
=========================

Scrapes the Sinokor E-Service schedule LIST TABLE (#tblSchedule) for each
Loading-port -> Discharging-port ROUTE and saves one .xlsx per route into
./schedule-daily/. The site only returns sailings when BOTH the loading port
(POL) and the discharging port (POD) are set before clicking Search — so this
scraper takes routes, not single ports.

Runs headless by default so it can be scheduled without opening a browser.

HOW TO RUN
----------
    pip install playwright openpyxl
    python -m playwright install chromium
    python daily_schedule_scraper.py --routes "LAEM CHABANG->BUSAN" "BANGKOK->BUSAN"

Route syntax:  "<POL name>-><POD name>"   e.g.  "LAEM CHABANG->BUSAN"
The names are matched via the site's own jQuery-UI autocomplete, so partial
matches like "LAEM" or "BUSAN" also work — but full names are less ambiguous.

Optional flags:
    --outdir schedule-daily     folder to save (default: ./schedule-daily)
    --headed                    show the browser (debug)
    --timeout 45                per-step timeout in seconds (default: 45)

WHAT IT DOES per route
----------------------
    1. Open https://ebiz.sinokor.co.kr/Schedule (fresh context, real Chrome UA)
    2. Type POL into #searchPol, click jQuery-UI autocomplete suggestion
    3. Type POD into #searchPod, click autocomplete suggestion
    4. Click #btnSearch
    5. Click the List toggle:  a[data-toggle='CalendarList'][data-title='L']
    6. Wait for #tblSchedule to render
    7. Read every <tr> from #tblSchedule tbody, extract cell text
    8. Save as <POL>_to_<POD>_<YYYY-MM-DD>.xlsx (well-formed xlsx via openpyxl)

If any step times out, the error is logged, a screenshot is saved into
./schedule-daily/_errors/, and the script moves on to the next route.

NO AI / NO API CALLS at runtime. Pure Playwright + openpyxl.

MAINTENANCE WARNING
-------------------
Selectors verified against the live DOM on 2026-09-06:
    #searchPol, #searchPod, ul.ui-autocomplete li.ui-menu-item, #btnSearch,
    a[data-toggle='CalendarList'][data-title='L'], #tblSchedule,
    EXPECTED_HEADERS below.
If Sinokor redesigns the Schedule page these will silently break. When that
happens, re-run a browser session with an AI assistant to re-inspect the DOM
and update the SELECTORS block below. Do NOT guess — verify against the real
site.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import (
    Page,
    TimeoutError as PWTimeoutError,
    sync_playwright,
)

# ------------------------------------------------------------------ SELECTORS
# All output goes under <script folder>/output/ by default so the script
# behaves the same no matter which working directory it was launched from.
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output" / "schedule-daily"

URL = "https://ebiz.sinokor.co.kr/Schedule"
SEL_POL_INPUT = "#searchPol"
SEL_POD_INPUT = "#searchPod"
SEL_AUTOCOMPLETE_ITEM = "ul.ui-autocomplete:visible li.ui-menu-item"
SEL_SEARCH_BUTTON = "#btnSearch"
SEL_LIST_TOGGLE = "a[data-toggle='CalendarList'][data-title='L']"
SEL_DATA_TABLE = "#tblSchedule"

# Column order shown by #tblSchedule thead (2026-09-06).
EXPECTED_HEADERS = [
    "T/S", "Service", "Vessel/Voyage", "Departure", "Arrival",
    "ETD", "ETA", "T/T", "DOCU Closing", "CNTR Closing", "Booking",
]

# The site returns an ASP.NET runtime error page to browsers whose UA reports
# "HeadlessChrome". A normal desktop Chrome UA avoids that.
REAL_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/128.0.0.0 Safari/537.36"
)
# ----------------------------------------------------------------------------


@dataclass
class Route:
    pol: str
    pod: str

    @classmethod
    def parse(cls, s: str) -> "Route":
        if "->" not in s:
            raise ValueError(
                f'route "{s}" must be "<POL>-><POD>", e.g. "LAEM CHABANG->BUSAN"'
            )
        pol, pod = [x.strip() for x in s.split("->", 1)]
        if not pol or not pod:
            raise ValueError(f'route "{s}" has empty POL or POD')
        return cls(pol=pol, pod=pod)

    def slug(self) -> str:
        return f"{_sanitize(self.pol)}_to_{_sanitize(self.pod)}"


def _sanitize(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_").upper()


def _fill_port(page: Page, selector: str, port: str, timeout_ms: int) -> None:
    """Type port into POL/POD, then click the first autocomplete suggestion."""
    box = page.locator(selector)
    box.click(timeout=timeout_ms)
    box.fill("")
    box.type(port, delay=60)
    page.wait_for_selector(SEL_AUTOCOMPLETE_ITEM, timeout=timeout_ms)
    page.locator(SEL_AUTOCOMPLETE_ITEM).first.click(timeout=timeout_ms)
    # Blur so the autocomplete dropdown for the next field can appear.
    page.evaluate(f"document.querySelector({selector!r}).blur()")
    page.wait_for_timeout(300)


def _read_table(page: Page) -> tuple[list[str], list[list[str]]]:
    """Return (headers, rows) from #tblSchedule. rows == [] if table shows
    the DataTables 'No data available' placeholder."""
    data = page.evaluate(
        """
        () => {
            const t = document.querySelector('#tblSchedule');
            if (!t) return null;
            const headers = Array.from(t.querySelectorAll('thead th'))
                .map(th => (th.innerText || '').trim());
            const rows = Array.from(t.querySelectorAll('tbody tr')).map(tr =>
                Array.from(tr.cells).map(td => (td.innerText || '').trim())
            );
            return {headers, rows};
        }
        """
    )
    if not data:
        return [], []
    headers = data["headers"]
    rows = data["rows"]
    if len(rows) == 1 and len(rows[0]) == 1 and "no data" in rows[0][0].lower():
        rows = []
    return headers, rows


def _write_xlsx(target: Path, route: Route, headers: list[str], rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(headers)
    fill = PatternFill("solid", fgColor="FF305496")
    font = Font(bold=True, color="FFFFFFFF")
    for cell in ws[1]:
        cell.fill, cell.font = fill, font
    for r in rows:
        padded = (r + [""] * len(headers))[: len(headers)]
        ws.append(padded)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)
    ws.freeze_panes = "A2"
    # Provenance sheet — the aggregator (and humans) don't need to parse the
    # filename to know POL/POD/snapshot.
    meta = wb.create_sheet("Meta")
    meta.append(["POL", route.pol])
    meta.append(["POD", route.pod])
    meta.append(["Snapshot", date.today().isoformat()])
    meta.append(["Source", URL])
    wb.save(target)


def scrape_one_route(page: Page, route: Route, outdir: Path, timeout_s: int) -> Path | None:
    step = "load-page"
    t_ms = timeout_s * 1000
    try:
        page.goto(URL, wait_until="domcontentloaded", timeout=t_ms)
        try:
            page.wait_for_load_state("networkidle", timeout=15_000)
        except PWTimeoutError:
            pass

        step = "fill-pol"
        _fill_port(page, SEL_POL_INPUT, route.pol, t_ms)

        step = "fill-pod"
        _fill_port(page, SEL_POD_INPUT, route.pod, t_ms)

        step = "click-search"
        page.locator(SEL_SEARCH_BUTTON).click(timeout=t_ms)
        try:
            page.wait_for_load_state("networkidle", timeout=t_ms)
        except PWTimeoutError:
            pass

        step = "toggle-list-view"
        page.locator(SEL_LIST_TOGGLE).click(timeout=t_ms)

        step = "wait-table"
        page.wait_for_selector(f"{SEL_DATA_TABLE} tbody tr", timeout=t_ms)
        page.wait_for_timeout(1200)  # let DataTables paint every cell

        step = "read-table"
        headers, rows = _read_table(page)
        if not headers:
            raise RuntimeError("#tblSchedule not present or thead empty")
        if headers != EXPECTED_HEADERS:
            print(
                f"[warn] {route.pol}->{route.pod}: table headers differ from expected: {headers!r}",
                file=sys.stderr,
            )

        step = "save-file"
        target = outdir / f"{route.slug()}_{date.today():%Y-%m-%d}.xlsx"
        _write_xlsx(target, route, headers, rows)
        print(f"[OK]   {route.pol}->{route.pod}: {len(rows)} row(s) -> {target}")
        return target

    except Exception as ex:
        errdir = outdir / "_errors"
        errdir.mkdir(parents=True, exist_ok=True)
        shot = errdir / f"{route.slug()}_{date.today():%Y-%m-%d}_{step}.png"
        try:
            page.screenshot(path=str(shot), full_page=True)
        except Exception:
            pass
        print(
            f"[FAIL] {route.pol}->{route.pod}: step={step!r} "
            f"error={type(ex).__name__}: {ex} — screenshot={shot}",
            file=sys.stderr,
        )
        return None


def run(routes: list[Route], outdir: Path, headless: bool, timeout_s: int) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        try:
            for route in routes:
                ctx = browser.new_context(
                    accept_downloads=False,
                    viewport={"width": 1440, "height": 900},
                    user_agent=REAL_UA,
                    locale="en-US",
                    extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
                )
                page = ctx.new_page()
                scrape_one_route(page, route, outdir, timeout_s)
                ctx.close()
        finally:
            browser.close()


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Scrape Sinokor Schedule list table per (POL,POD) route."
    )
    ap.add_argument(
        "--routes", nargs="+", required=True,
        help='Routes in "POL->POD" form, e.g. --routes "LAEM CHABANG->BUSAN" "BANGKOK->BUSAN"',
    )
    ap.add_argument("--outdir", default=str(DEFAULT_OUTPUT_DIR),
                    help="Directory to save output files "
                         "(default: <script-dir>/output/schedule-daily)")
    ap.add_argument("--headed", action="store_true",
                    help="Show the browser window (default: headless).")
    ap.add_argument("--timeout", type=int, default=45,
                    help="Per-step timeout in seconds (default: 45).")
    args = ap.parse_args()

    routes = [Route.parse(s) for s in args.routes]

    run(routes=routes, outdir=Path(args.outdir),
        headless=not args.headed, timeout_s=args.timeout)


if __name__ == "__main__":
    main()
