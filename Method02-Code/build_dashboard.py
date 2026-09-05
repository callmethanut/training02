"""
build_dashboard.py
==================

Reads every *.xlsx file that daily_schedule_scraper.py has dropped into
./schedule-daily/ and produces Dashboard.xlsx with:

    Raw     — every schedule row from every route, prefixed with POL, POD and
              the snapshot date.
    Weekly  — one row per ISO week (Mon start), one column per route + Total.
              Counts are LIVE Excel COUNTIFS formulas pointing back into Raw —
              editing Raw updates Weekly on the next recalculation. No values
              hardcoded.
    A line chart of weekly totals per route is embedded on the Weekly sheet.

HOW TO RUN
----------
    pip install openpyxl
    python build_dashboard.py
    # or:
    python build_dashboard.py --indir schedule-daily --out Dashboard.xlsx

Assumes source files were produced by daily_schedule_scraper.py — each file
has a "Schedule" sheet with the 11-column layout below and a "Meta" sheet
that carries POL / POD / Snapshot / Source.

NO AI / NO API CALLS at runtime. Pure openpyxl.

MAINTENANCE WARNING
-------------------
If daily_schedule_scraper.py changes the sheet layout (columns, sheet names,
meta keys), update SOURCE_COLUMNS / RAW_HEADERS / _read_source() below.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Default I/O lives under <script folder>/output/ so this script runs the
# same from any working directory.
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INDIR = SCRIPT_DIR / "output" / "schedule-daily"
DEFAULT_OUT = SCRIPT_DIR / "output" / "Dashboard.xlsx"

SOURCE_COLUMNS = [
    "T/S", "Service", "Vessel/Voyage", "Departure", "Arrival",
    "ETD", "ETA", "T/T", "DOCU Closing", "CNTR Closing", "Booking",
]
RAW_HEADERS = ["Route", "POL", "POD", "SnapshotDate"] + SOURCE_COLUMNS

HEADER_FILL = PatternFill("solid", fgColor="FF305496")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="FFE7E6E6")

_FNAME_RE = re.compile(r"^(?P<slug>.+?)_(?P<date>\d{4}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE)


@dataclass
class SourceFile:
    path: Path
    pol: str
    pod: str
    snapshot: date

    @property
    def route(self) -> str:
        return f"{self.pol}->{self.pod}"

    @property
    def route_slug(self) -> str:
        return f"{_slug(self.pol)}_to_{_slug(self.pod)}"


def _slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").upper()


def _read_source(path: Path) -> tuple[str, str, date, list[list[str]]] | None:
    """Return (pol, pod, snapshot, data_rows) from a scraper output file."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as ex:
        print(f"[read-fail] {path.name}: {ex}", file=sys.stderr)
        return None

    if "Meta" not in wb.sheetnames or "Schedule" not in wb.sheetnames:
        print(f"[skip] {path.name}: missing Meta or Schedule sheet", file=sys.stderr)
        return None

    meta = {str(row[0]).strip(): (str(row[1]).strip() if row[1] is not None else "")
            for row in wb["Meta"].iter_rows(values_only=True) if row and row[0] is not None}
    pol = meta.get("POL", "")
    pod = meta.get("POD", "")
    snap_str = meta.get("Snapshot", "")
    try:
        snap = datetime.strptime(snap_str, "%Y-%m-%d").date()
    except ValueError:
        # Fall back to filename date
        m = _FNAME_RE.match(path.name)
        snap = datetime.strptime(m.group("date"), "%Y-%m-%d").date() if m else date.today()

    # Read Schedule rows (skip header row 1).
    rows: list[list[str]] = []
    for r in wb["Schedule"].iter_rows(min_row=2, values_only=True):
        if r is None or all((c is None or str(c).strip() == "") for c in r):
            continue
        rows.append([("" if c is None else str(c)) for c in r])

    return pol, pod, snap, rows


def discover_sources(indir: Path) -> list[SourceFile]:
    out: list[SourceFile] = []
    for p in sorted(indir.glob("*.xlsx")):
        parsed = _read_source(p)
        if parsed is None:
            continue
        pol, pod, snap, _rows = parsed
        if not pol or not pod:
            print(f"[skip] {p.name}: empty POL/POD in Meta", file=sys.stderr)
            continue
        out.append(SourceFile(path=p, pol=pol, pod=pod, snapshot=snap))
    return out


def parse_etd(cell: str) -> date | None:
    if not cell:
        return None
    cell = cell.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(cell.split()[0], fmt).date()
        except ValueError:
            continue
    return None


def week_start(d: date) -> date:
    return d - timedelta(days=d.weekday())


def build_workbook(sources: list[SourceFile], out_path: Path) -> None:
    wb = Workbook()

    # Raw ------------------------------------------------------------------
    raw = wb.active
    raw.title = "Raw"
    raw.append(RAW_HEADERS)
    for cell in raw[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    row_count = 0
    routes_seen: list[str] = []
    for src in sources:
        parsed = _read_source(src.path)
        if parsed is None:
            continue
        _pol, _pod, _snap, rows = parsed
        for r in rows:
            padded = (r + [""] * len(SOURCE_COLUMNS))[: len(SOURCE_COLUMNS)]
            raw.append([src.route, src.pol, src.pod, src.snapshot.isoformat(), *padded])
            row_count += 1
        if src.route not in routes_seen:
            routes_seen.append(src.route)

    raw.freeze_panes = "A2"
    for i, h in enumerate(RAW_HEADERS, start=1):
        raw.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    # Weekly ---------------------------------------------------------------
    weekly = wb.create_sheet("Weekly")
    weekly.append(["Week (Mon)", *routes_seen, "Total"])
    for cell in weekly[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    weeks = _collect_weeks(raw, row_count)
    route_col_letter = "A"
    etd_col_letter = get_column_letter(RAW_HEADERS.index("ETD") + 1)
    ports_range = f"Raw!${route_col_letter}$2:${route_col_letter}${row_count + 1}"
    etds_range = f"Raw!${etd_col_letter}$2:${etd_col_letter}${row_count + 1}"

    for w in weeks:
        we = w + timedelta(days=6)
        row_vals: list = [w]
        for route in routes_seen:
            # ETD is stored as ISO text — lexical compare == chronological compare.
            f = (
                f'=IF({row_count}=0,0,'
                f'COUNTIFS({ports_range},"{route}",'
                f'{etds_range},">={w.isoformat()}",'
                f'{etds_range},"<={we.isoformat()}"))'
            )
            row_vals.append(f)
        weekly.append(row_vals)
        r = weekly.max_row
        first = get_column_letter(2)
        last = get_column_letter(1 + len(routes_seen))
        weekly.cell(row=r, column=len(routes_seen) + 2,
                    value=f"=SUM({first}{r}:{last}{r})")

    total_col_idx = len(routes_seen) + 2
    for r in range(2, weekly.max_row + 1):
        weekly.cell(row=r, column=total_col_idx).fill = TOTAL_FILL
    for i in range(1, total_col_idx + 1):
        weekly.column_dimensions[get_column_letter(i)].width = max(14, len(routes_seen[i - 2]) + 2) if 2 <= i <= 1 + len(routes_seen) else 14
    weekly.freeze_panes = "B2"
    for cell in weekly["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(horizontal="left")

    # Chart ----------------------------------------------------------------
    if weekly.max_row > 1 and routes_seen:
        chart = LineChart()
        chart.title = "Weekly sailings per route"
        chart.y_axis.title = "Sailings"
        chart.x_axis.title = "Week starting"
        chart.height = 12
        chart.width = 24
        data = Reference(
            weekly,
            min_col=2, max_col=1 + len(routes_seen),
            min_row=1, max_row=weekly.max_row,
        )
        cats = Reference(weekly, min_col=1, min_row=2, max_row=weekly.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        weekly.add_chart(chart, f"{get_column_letter(total_col_idx + 2)}2")

    wb.save(out_path)
    print(f"[OK]  Wrote {out_path}  (raw rows: {row_count}, weeks: {len(weeks)}, routes: {len(routes_seen)})")


def _collect_weeks(raw_sheet, row_count: int) -> list[date]:
    weeks: set[date] = set()
    etd_col = RAW_HEADERS.index("ETD") + 1
    for r in range(2, row_count + 2):
        v = raw_sheet.cell(row=r, column=etd_col).value
        if v is None:
            continue
        d = parse_etd(str(v))
        if d:
            weeks.add(week_start(d))
    return sorted(weeks)


def main() -> None:
    ap = argparse.ArgumentParser(description="Aggregate schedule-daily/*.xlsx into Dashboard.xlsx")
    ap.add_argument("--indir", default=str(DEFAULT_INDIR),
                    help="Folder with per-route Excel files "
                         "(default: <script-dir>/output/schedule-daily)")
    ap.add_argument("--out", default=str(DEFAULT_OUT),
                    help="Output workbook path "
                         "(default: <script-dir>/output/Dashboard.xlsx)")
    args = ap.parse_args()

    indir = Path(args.indir)
    if not indir.is_dir():
        print(f"[fatal] input folder not found: {indir}", file=sys.stderr)
        sys.exit(2)

    sources = discover_sources(indir)
    if not sources:
        print("[fatal] no usable .xlsx files found in", indir, file=sys.stderr)
        sys.exit(2)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(sources, out_path)


if __name__ == "__main__":
    main()
